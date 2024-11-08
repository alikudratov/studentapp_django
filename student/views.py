from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Sum, Min, Max
from django.contrib import messages
from django.http import JsonResponse, FileResponse
from django.contrib.auth.decorators import login_required
from django.views.generic import ListView, UpdateView
from django.urls import reverse_lazy
from django.db.models import Count
#from qr_code.qr_code_views import make_qr_code
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User

import openpyxl
from django.http import HttpResponse

from .models import *
from .forms import *

def pagevisits(request):
    newvisit = PageVisits()
    newvisit.ip_address = request.META['REMOTE_ADDR']
    newvisit.path = request.path
    newvisit.username = request.user
    newvisit.save()

def appsetting():
    settings = {
        'code100' : Sozlamalar.objects.get(kodi=100),
        'code200' : Sozlamalar.objects.get(kodi=200),
        'code300' : Sozlamalar.objects.get(kodi=300),
    }
    return settings

def statistics():
    guruh_c = Talaba.objects.values('guruh__guruh_raqami').annotate(count=Count('id'))
    jinsi_c = Talaba.objects.values('jinsi__jins').annotate(count=Count('id'))
    viloyat_c = Talaba.objects.values('viloyat__viloyat_nomi').annotate(count=Count('id'))
    user_c = Talaba.objects.values('created_by__username').annotate(count=Count('id'))

    context = {
        'guruh_c' : guruh_c,
        'jinsi_c' : jinsi_c,
        'viloyat_c' : viloyat_c,
        'user_c' : user_c
    }

    return context

def index(request):

    pagevisits(request)

    user_reg_form = UserRegForm(request.POST or None)

    if user_reg_form.is_valid():
        username = user_reg_form.cleaned_data["username"]
        password = user_reg_form.cleaned_data["password"]
        new_user = User.objects.create_user(username = username, password = password)
        return redirect('login')

    search_form = SearchForm(request.GET or None)

    if search_form.is_valid():
        search = search_form.cleaned_data["search"]
        students = Talaba.objects.filter(familiya__contains = search).order_by('-id')
    else:
        students = Talaba.objects.all().order_by('-id')

    count = Talaba.objects.count()

    form = MurojaatForm(request.POST or None)

    if form.is_valid():
        form.save()
        messages.success(request, 'Murojaat muvaffaqiyatli yuborildi!')

    context = {
        'count' : count, 
        'students' : students, 
        'form' : form, 
        'search_form' : search_form,
        'statistika': statistics(),
        'settings' : appsetting(),
        'user_reg_form' : user_reg_form,
        }
    
    return render(request, 'index.html', context)

def talabapage(request, id):

    pagevisits(request)
    
    talaba1 = Talaba.objects.get(id=id)

    comments_list = Comments.objects.filter(talaba_id = id, published = True).order_by("-id")
    
    comment_form = CommentsForm(request.POST or None, initial={'talaba':id})

    if comment_form.is_valid():
        comment_form.save()
        redirect(request.path)

    return render(request, 'test.html', 
                  {'talaba': talaba1, 
                   'comment_form':comment_form, 
                   'izohlar':comments_list,
                   'statistika': statistics()})

@login_required(login_url="/login")
def fanlarpage(request):
    
    pagevisits(request)

    fanlar1 = Fanlar.objects.all()

    oxirgi_talaba = Talaba.objects.last()

    total_price = Fanlar.objects.aggregate(Sum('kitob_narxi'))

    return render(request, 'fanlar.html', {'fanlar': fanlar1, 'oxirgi_talaba':oxirgi_talaba, 'total_price':total_price})

def webservice(request, id):

    pagevisits(request)

    talaba1 = Talaba.objects.filter(id=id).values()
    data = list(talaba1)
    return JsonResponse(data, safe=False)

def download_file(request, filename):

    pagevisits(request)

    return FileResponse(open('data/' + filename, 'rb'), as_attachment=False)

class TalabaListView(ListView):
    model = Talaba
    template_name = 'talabalist.html'
    context_object_name = 'object_list'
    paginate_by = 2

class TalabaUpdateView(UpdateView):
    model = Talaba
    template_name = 'talabalist.html'
    fields = '__all__'
    success_url = reverse_lazy('index')

def datatable(request):
    students = Talaba.objects.all()
    return render(request, 'slist.html', {'students' : students})

def generate_pdf(request):

    pagevisits(request)

    # Retrieve HTML content from a template or generate it dynamically
    #qr_img = make_qr_code('https://my.gov.uz/uz', image_factory=None)
    talaba_one = Talaba.objects.all()
    template = get_template('pdf.html')
    html = template.render({'allstudents': talaba_one})

    # Create a PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'filename="generated_pdf.pdf"'

    # Generate PDF from HTML
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err:
        return HttpResponse('Error generating PDF: %s' % pisa_status.err)

    return response

def update_talaba(request):

    pagevisits(request)

    data = Talaba.objects.get(pk=2)
    talabaform = TalabaUpdateForm(instance=data)
    return render(request,'update.html', {'talabaform':talabaform})

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('index')
        else:
            return render(request, 'login.html', {'error': 'Invalid username or password'})
    return render(request, 'login.html')

def logout_view(request):
    logout(request)
    return redirect('login')

def talaba_related(request):
    guruhlar = Guruh.objects.all()
    jinslar = Jinsi.objects.all()
    viloyat = Viloyatlar.objects.all()
    return render(request, 
                  'talaba_related.html',
                  {'guruhlar' : guruhlar,
                  'jinslar' : jinslar,
                  'viloyat' : viloyat})

def export_excel(request, stir):
    x = stir
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="mydata_{x}.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'My Data'

        # Write header row
    header = ['hudud', 'tuman', 'maktab', 'sinf', 'fish', 'id']
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Write data rows
    sinflar = ['7-', '8-', '9-', '10-']
    for y in sinflar:
        queryset = Maktablar.objects.filter(stir = x, sinf__startswith = y).values_list('hudud', 'tuman', 'maktab', 'sinf', 'fish', 'id')
        for row_num, row in enumerate(queryset, 1):
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num+1, column=col_num)
                cell.value = cell_value

    workbook.save(response)

    return response